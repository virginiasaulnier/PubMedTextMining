import logging
import os
import xlrd
import xlwt
import numpy

from openpyxl import Workbook
from openpyxl import load_workbook



from Bio import Entrez
Entrez.email = "virginia.saulnier@gmail.com"

#input file
import csv

with open('names.csv', 'w') as csvfile:
    fieldnames = ['first_name', 'last_name']
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerow({'first_name': 'Baked', 'last_name': 'Beans'})
    writer.writerow({'first_name': 'Lovely', 'last_name': 'Spam'})
    writer.writerow({'first_name': 'Wonderful', 'last_name': 'Spam'})
#wb = load_workbook("/Users/virginiasaulnier/Documents/emaadversereactions.xlsx")
workbook= xlrd.open_workbook("/Users/virginiasaulnier/Documents/emaadversereactions.xlsx")
#print(wb.sheetnames)

sheet = workbook.sheet_by_index(0)
i=1
workbook2 = xlwt.Workbook()
#workbook2.save('newadversereactions.xls')
sheet2 = workbook2.add_sheet('Sheet_1')

while sheet.cell(i,3).value != xlrd.empty_cell.value:
    effect = sheet.cell(i,3).value
    handle = Entrez.esearch(db="pubmed",retmax=10, term="(tumour necrosis factor[All Fields]"
                                            "OR tumor necrosis factor-alpha[MeSH Terms] OR (tumor[All Fields])"
                                            "AND necrosis[All Fields] AND factor-alpha[All Fields]) OR "
                                            "tumor necrosis factor-alpha[All Fields] OR (tumor[All Fields] AND "
                                            "necrosis[All Fields] AND factor[All Fields]) OR "
                                            "tumor necrosis factor[All Fields]) AND "
                                            "inhibitor[All Fields] AND "+str(effect)+"[All Fields]))")
    record = Entrez.read(handle)
    #sheet2.write(i,4, "+str(effect)+")

    record["Count"]
    record["IdList"]
    print("The first 10 are\n{}".format(record['IdList']))
    PubMedIDs = dict({"+str(effect)+": ".format(record['IdList']"})
    i=i+1
    print(i)
    break



print(PubMedIDs)
